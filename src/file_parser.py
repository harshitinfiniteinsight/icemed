"""
Excel File Parser - Reads ICE export files and parses to Encounter objects
"""

from openpyxl import load_workbook
from typing import List, Tuple
from .models import Encounter
import logging

logger = logging.getLogger(__name__)


class ParseError:
    """Represents a parsing error"""
    def __init__(self, row_num: int, field: str, message: str):
        self.row_num = row_num
        self.field = field
        self.message = message
    
    def __str__(self):
        return f"Row {self.row_num}, Field '{self.field}': {self.message}"


class ExcelFileParser:
    """Parser for ICE Excel export files"""
    
    REQUIRED_COLUMNS = [
        "Patient Name", "DOB", "Date of Service", "Type of Care", "Type of Visit",
        "Facility", "Servicing Provider", "Supervising Provider"
    ]
    
    def __init__(self, config: dict = None):
        """Initialize parser with configuration"""
        self.config = config or {}
        self.sheet_name = self.config.get("sheetName", "Sheet1")
    
    def parse_file(self, file_path: str) -> Tuple[List[Encounter], List[ParseError]]:
        """
        Parse Excel file and return list of encounters and errors
        
        Returns:
            Tuple of (encounters, errors)
        """
        encounters = []
        errors = []
        
        try:
            # Load workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheet (try by name, fallback to first sheet)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
            
            # Get header row
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                logger.error("Excel file is empty")
                return encounters, errors
            
            header = rows[0]
            
            # Validate required columns
            missing_cols = self.validate_columns(header)
            if missing_cols:
                error_msg = f"Missing required columns: {', '.join(missing_cols)}"
                logger.error(error_msg)
                errors.append(ParseError(1, "Header", error_msg))
                return encounters, errors
            
            # Create column index mapping
            col_map = {col: idx for idx, col in enumerate(header)}
            
            # Parse data rows
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    encounter = self.parse_row(row, col_map)
                    encounters.append(encounter)
                except Exception as e:
                    logger.warning(f"Error parsing row {row_num}: {e}")
                    errors.append(ParseError(row_num, "Row", str(e)))
            
            logger.info(f"Parsed {len(encounters)} encounters with {len(errors)} errors")
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            errors.append(ParseError(0, "File", str(e)))
        
        return encounters, errors
    
    def validate_columns(self, header: tuple) -> List[str]:
        """
        Validate that required columns exist
        Returns list of missing columns
        """
        header_list = [str(col) for col in header if col]
        missing = []
        
        for required_col in self.REQUIRED_COLUMNS:
            if required_col not in header_list:
                missing.append(required_col)
        
        return missing
    
    def parse_row(self, row: tuple, col_map: dict) -> Encounter:
        """Parse a single row into an Encounter object"""
        
        def get_value(col_name: str) -> str:
            """Get value from row by column name"""
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row):
                value = row[idx]
                return str(value) if value is not None else ""
            return ""
        
        encounter = Encounter(
            patient_name=get_value("Patient Name"),
            dob=get_value("DOB"),
            date_of_service=get_value("Date of Service"),
            type_of_care=get_value("Type of Care"),
            type_of_visit=get_value("Type of Visit"),
            facility=get_value("Facility"),
            room=get_value("Room"),
            assessment=get_value("Assessment"),
            cpt=get_value("CPT"),
            chief_complaint=get_value("Chief Complaint"),
            visit_type=get_value("Visit Type"),
            servicing_provider=get_value("Servicing Provider"),
            supervising_provider=get_value("Supervising Provider"),
            time=get_value("Time"),
            code_status=get_value("Code Status"),
            observation=get_value("Observation"),
            encounter_status=get_value("Encounter Status"),
            status_aux=get_value("Status Aux"),
            export_date=get_value("Export Date")
        )
        
        return encounter
