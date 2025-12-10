"""
Master Missing Manager - Manages the historical Master Missing file
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from typing import Dict, List
from .models import Encounter, BillingResult, MasterMissingRecord
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


class MasterMissingManager:
    """Manages Master Missing file (historical ledger of incomplete encounters)"""
    
    def __init__(self, config: dict = None):
        """Initialize manager with configuration"""
        self.config = config or {}
        self.folder_path = self.config.get("folderPath", "data/output")
        self.file_pattern = self.config.get("fileNamePattern", "Master Missing to {date}.xlsx")
    
    def load_previous_file(self, file_path: str = None) -> Dict[str, MasterMissingRecord]:
        """
        Load previous Master Missing file
        
        Args:
            file_path: Specific file path, or None to find latest
        
        Returns:
            Dictionary mapping encounter_key to MasterMissingRecord
        """
        if not file_path:
            file_path = self._find_latest_master_missing_file()
        
        if not file_path or not os.path.exists(file_path):
            logger.info("No previous Master Missing file found, starting fresh")
            return {}
        
        records = {}
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return records
            
            header = rows[0]
            col_map = {col: idx for idx, col in enumerate(header)}
            
            for row in rows[1:]:
                try:
                    record = self._parse_master_missing_row(row, col_map)
                    if record.encounter_key:
                        records[record.encounter_key] = record
                except Exception as e:
                    logger.warning(f"Error parsing Master Missing row: {e}")
            
            logger.info(f"Loaded {len(records)} records from previous Master Missing file")
            
        except Exception as e:
            logger.error(f"Error loading Master Missing file: {e}")
        
        return records
    
    def update_with_results(self, previous_records: Dict[str, MasterMissingRecord],
                           encounters: List[Encounter], 
                           billing_results: List[BillingResult],
                           execution_date: str) -> Dict[str, MasterMissingRecord]:
        """
        Update Master Missing records based on current billing results
        
        Logic:
        - If billing failed: add new or update existing record
        - If billing succeeded: remove from Master Missing (if exists)
        
        Returns:
            Updated dictionary of Master Missing records
        """
        updated_records = previous_records.copy()
        
        # Create billing results map
        billing_map = {result.encounter_key: result for result in billing_results}
        
        added = 0
        updated = 0
        removed = 0
        
        for encounter in encounters:
            key = encounter.generate_key()
            result = billing_map.get(key)
            
            if result and result.success:
                # Billing succeeded - remove from Master Missing if exists
                if key in updated_records:
                    del updated_records[key]
                    removed += 1
                    logger.debug(f"Removed from Master Missing: {encounter.patient_name}")
            else:
                # Billing failed - add or update Master Missing record
                reason = result.reason if result else "Unknown Error"
                
                if key in updated_records:
                    # Update existing record
                    updated_records[key].last_attempt_to_process = execution_date
                    updated_records[key].reason_for_not_billed = reason
                    updated += 1
                    logger.debug(f"Updated Master Missing: {encounter.patient_name}")
                else:
                    # Add new record
                    record = MasterMissingRecord.from_encounter(encounter, reason, execution_date)
                    updated_records[key] = record
                    added += 1
                    logger.debug(f"Added to Master Missing: {encounter.patient_name}")
        
        logger.info(f"Master Missing updates: Added={added}, Updated={updated}, Removed={removed}")
        
        return updated_records, {"added": added, "updated": updated, "removed": removed}
    
    def write_file(self, records: Dict[str, MasterMissingRecord], 
                   output_path: str, execution_date: str = None) -> None:
        """
        Write Master Missing file
        
        Args:
            records: Dictionary of Master Missing records
            output_path: Path to output file
            execution_date: Execution date (for filename)
        """
        if not execution_date:
            execution_date = datetime.now().strftime("%m-%d-%Y")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Headers
        headers = [
            "Patient Name", "DOB", "Date of Service", "Type of Care", 
            "Type of Visit", "Facility", "Last Attempt to Process", 
            "Billed", "Reason for not billed"
        ]
        
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Sort records by Date of Service
        sorted_records = sorted(
            records.values(), 
            key=lambda r: r.date_of_service
        )
        
        # Write data rows
        for record in sorted_records:
            row = [
                record.patient_name,
                record.dob,
                record.date_of_service,
                record.type_of_care,
                record.type_of_visit,
                record.facility,
                record.last_attempt_to_process,
                record.billed,
                record.reason_for_not_billed
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file (use /tmp if original path is read-only)
        try:
            wb.save(output_path)
        except (OSError, PermissionError):
            # If we can't write to the original path (read-only filesystem),
            # save to /tmp instead
            import tempfile
            import os
            temp_dir = tempfile.gettempdir()
            filename = os.path.basename(output_path)
            temp_path = os.path.join(temp_dir, filename)
            wb.save(temp_path)
            logger.warning(f"Cannot write to {output_path}, saved to {temp_path} instead")
            # Update output_path for caller
            output_path = temp_path
        logger.info(f"Saved Master Missing file: {output_path} ({len(records)} records)")
    
    def _find_latest_master_missing_file(self) -> str:
        """Find the most recent Master Missing file in the folder"""
        if not os.path.exists(self.folder_path):
            return None
        
        master_files = [
            f for f in os.listdir(self.folder_path)
            if f.startswith("Master Missing") and f.endswith(".xlsx")
        ]
        
        if not master_files:
            return None
        
        # Sort by modification time, return latest
        master_files.sort(
            key=lambda f: os.path.getmtime(os.path.join(self.folder_path, f)),
            reverse=True
        )
        
        return os.path.join(self.folder_path, master_files[0])
    
    def _parse_master_missing_row(self, row: tuple, col_map: dict) -> MasterMissingRecord:
        """Parse a row from Master Missing file"""
        def get_value(col_name: str) -> str:
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row):
                value = row[idx]
                return str(value) if value is not None else ""
            return ""
        
        # Create a temporary encounter to generate the key
        from .models import Encounter
        temp_encounter = Encounter(
            patient_name=get_value("Patient Name"),
            dob=get_value("DOB"),
            date_of_service=get_value("Date of Service"),
            type_of_care=get_value("Type of Care"),
            type_of_visit=get_value("Type of Visit"),
            facility=get_value("Facility"),
            room="",  # Not in Master Missing
            assessment="",  # Not in Master Missing
            cpt="",  # Not in Master Missing
            chief_complaint="",
            visit_type="",
            servicing_provider="",
            supervising_provider="",
            time="",
            code_status="",
            observation="",
            encounter_status="",
            status_aux="",
            export_date=""
        )
        
        record = MasterMissingRecord(
            patient_name=get_value("Patient Name"),
            dob=get_value("DOB"),
            date_of_service=get_value("Date of Service"),
            type_of_care=get_value("Type of Care"),
            type_of_visit=get_value("Type of Visit"),
            facility=get_value("Facility"),
            last_attempt_to_process=get_value("Last Attempt to Process"),
            billed=get_value("Billed"),
            reason_for_not_billed=get_value("Reason for not billed"),
            encounter_key=temp_encounter.generate_key()
        )
        
        return record
