"""
General Reconciliation Generator - Creates General Reconciliation Excel file
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from typing import List
from .models import Encounter, BillingResult, ReconciliationData
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class GeneralReconciliationGenerator:
    """Generates General Reconciliation Excel file with Data and Summary sheets"""
    
    def __init__(self, config: dict = None):
        """Initialize generator with configuration"""
        self.config = config or {}
        self.date_format = self.config.get("dateFormat", "MM-dd-yyyy")
    
    def generate(self, encounters: List[Encounter], billing_results: List[BillingResult], 
                 output_path: str, execution_date: str = None) -> str:
        """
        Generate General Reconciliation Excel file
        
        Args:
            encounters: List of encounters
            billing_results: List of billing results (same order as encounters)
            output_path: Path to output file
            execution_date: Execution date (defaults to today)
        """
        if not execution_date:
            execution_date = datetime.now().strftime("%m-%d-%Y")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Data sheet
        self._create_data_sheet(wb, encounters, billing_results)
        
        # Create Summary sheet
        self._create_summary_sheet(wb, encounters, billing_results)
        
        # Save file (use /tmp if original path is read-only)
        actual_output_path = output_path
        try:
            wb.save(output_path)
        except (OSError, PermissionError):
            # If we can't write to the original path (read-only filesystem),
            # save to /tmp instead
            import tempfile
            import os
            temp_dir = tempfile.gettempdir()
            filename = os.path.basename(output_path)
            temp_path = os.path.join(temp_dir, filename)
            wb.save(temp_path)
            logger.warning(f"Cannot write to {output_path}, saved to {temp_path} instead")
            # Update actual_output_path for caller
            actual_output_path = temp_path
        
        logger.info(f"Generated General Reconciliation file: {actual_output_path}")
        
        # Return the actual path where file was saved
        return actual_output_path
    
    def _create_data_sheet(self, wb: Workbook, encounters: List[Encounter], 
                          billing_results: List[BillingResult]) -> None:
        """Create Data sheet with all encounters and billing status"""
        ws = wb.create_sheet("Data", 0)
        
        # Headers
        headers = [
            "Patient Name", "DOB", "Date of Service", "Type of Care", "Type of Visit",
            "Facility", "Room", "Assessment", "CPT", "Chief Complaint", 
            "Visit Type", "Servicing Provider", "Supervising Provider", 
            "Time", "Code Status", "Observation", "Encounter Status", 
            "Status Aux", "Export Date", "Billed", "Reason for not billed"
        ]
        
        # Write header row with formatting
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Create billing results map
        billing_map = {result.encounter_key: result for result in billing_results}
        
        # Write data rows
        for encounter in encounters:
            key = encounter.generate_key()
            result = billing_map.get(key)
            
            billed = "Yes" if result and result.success else "No"
            reason = result.reason if result and not result.success else ""
            
            row = [
                encounter.patient_name,
                encounter.dob,
                encounter.date_of_service,
                encounter.type_of_care,
                encounter.type_of_visit,
                encounter.facility,
                encounter.room,
                encounter.assessment,
                encounter.cpt,
                encounter.chief_complaint,
                encounter.visit_type,
                encounter.servicing_provider,
                encounter.supervising_provider,
                encounter.time,
                encounter.code_status,
                encounter.observation,
                encounter.encounter_status,
                encounter.status_aux,
                encounter.export_date,
                billed,
                reason
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb: Workbook, encounters: List[Encounter], 
                             billing_results: List[BillingResult]) -> None:
        """Create Summary sheet with aggregated statistics"""
        ws = wb.create_sheet("Summary", 1)
        
        # Headers
        headers = ["Date", "Facility", "Provider", "Type of Care", "PRM Billing", "CPTs"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Create billing results map
        billing_map = {result.encounter_key: result for result in billing_results}
        
        # Aggregate data
        summary = self._aggregate_summary(encounters, billing_map)
        
        # Write summary rows
        for item in summary:
            row = [
                item["date"],
                item["facility"],
                item["provider"],
                item["type_of_care"],
                item["prm_billing"],
                item["cpts"]
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _aggregate_summary(self, encounters: List[Encounter], 
                          billing_map: dict) -> List[dict]:
        """
        Aggregate encounters by Date, Facility, Provider, Type of Care
        Only include successfully billed encounters
        """
        # Group by key
        groups = {}
        
        for encounter in encounters:
            key = encounter.generate_key()
            result = billing_map.get(key)
            
            # Only include successfully billed
            if not result or not result.success:
                continue
            
            # Create grouping key
            group_key = (
                encounter.date_of_service,
                encounter.facility,
                encounter.servicing_provider,
                encounter.type_of_care
            )
            
            if group_key not in groups:
                groups[group_key] = {
                    "date": encounter.date_of_service,
                    "facility": encounter.facility,
                    "provider": encounter.servicing_provider,
                    "type_of_care": encounter.type_of_care,
                    "prm_billing": 0,
                    "cpts": 0
                }
            
            groups[group_key]["prm_billing"] += 1
            groups[group_key]["cpts"] += 1  # Assuming 1 CPT per encounter
        
        # Convert to list and sort by date
        summary = list(groups.values())
        summary.sort(key=lambda x: x["date"])
        
        return summary
