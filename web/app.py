"""
Flask Web Application for ICE Reconciliation Mock System
"""

from flask import Flask, render_template, request, jsonify, send_file, session
import os
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
import sys

# Add parent directory to path to import src modules
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, PROJECT_ROOT)

from src.orchestrator import ReconciliationOrchestrator

# Initialize Flask with explicit paths for Vercel
template_dir = os.path.join(PROJECT_ROOT, 'web', 'templates')
static_dir = os.path.join(PROJECT_ROOT, 'web', 'static')

app = Flask(__name__, 
            template_folder=template_dir,
            static_folder=static_dir,
            static_url_path='/static')
app.secret_key = 'ice-reconciliation-secret-key-change-in-production'

# Configuration
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, 'data', 'input', 'uploads')
ALLOWED_EXTENSIONS = {'xlsx'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Create directories if they don't exist (may fail in read-only Vercel, that's OK)
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except Exception:
    pass  # May fail in serverless environment, that's OK

# Initialize orchestrator with absolute path to config
# Defer initialization to avoid read-only filesystem issues at import time
CONFIG_PATH = os.path.join(PROJECT_ROOT, 'config.json')
orchestrator = None

def get_orchestrator():
    """Lazy initialization of orchestrator"""
    global orchestrator
    if orchestrator is None:
        try:
            orchestrator = ReconciliationOrchestrator(CONFIG_PATH)
        except Exception as e:
            print(f"Warning: Failed to initialize orchestrator: {e}")
            # Return None - routes will handle this
    return orchestrator

# Store results in memory (in production, use Redis or database)
results_store = {}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_preview_data(reconciliation_file_path, max_rows=20):
    """Extract preview data from reconciliation file"""
    from openpyxl import load_workbook
    preview_data = []
    try:
        if not reconciliation_file_path or not os.path.exists(reconciliation_file_path):
            print(f"Warning: Reconciliation file not found: {reconciliation_file_path}")
            return preview_data
        
        wb = load_workbook(reconciliation_file_path, read_only=True, data_only=True)
        ws = wb['Data']
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        # Get first max_rows of data (skip header)
        for row in rows[1:max_rows+1]:
            if row:  # Skip empty rows
                preview_data.append(dict(zip(headers, row)))
        wb.close()
        print(f"Successfully loaded {len(preview_data)} preview rows from {reconciliation_file_path}")
    except Exception as e:
        print(f"Error reading preview data from {reconciliation_file_path}: {e}")
        import traceback
        traceback.print_exc()
    return preview_data


@app.route('/')
def index():
    """Upload page"""
    return render_template('index.html')


@app.route('/api/sample-files', methods=['GET'])
def get_sample_files():
    """Get list of available sample files"""
    sample_dir = os.path.join(PROJECT_ROOT, 'data', 'input')
    
    try:
        if not os.path.exists(sample_dir):
            return jsonify({
                'success': False,
                'error': f'Sample directory not found: {sample_dir}'
            }), 404
            
        files = [
            f for f in os.listdir(sample_dir)
            if f.startswith('sample_') and f.endswith('.xlsx')
        ]
        
        # Create file info with descriptions
        file_info = []
        descriptions = {
            'sample_complete.xlsx': 'All complete encounters (20 encounters)',
            'sample_missing_dx.xlsx': 'Missing DX codes (15 encounters, 10 missing DX)',
            'sample_missing_cpt.xlsx': 'Missing CPT codes (15 encounters, 8 missing CPT)',
            'sample_mixed.xlsx': 'Mixed scenarios (25 encounters)',
            'sample_multiple_dates.xlsx': 'Multiple dates (30 encounters, 3 dates)',
            'sample_large.xlsx': 'Large file (150 encounters)'
        }
        
        for filename in sorted(files):
            file_info.append({
                'filename': filename,
                'description': descriptions.get(filename, filename)
            })
        
        return jsonify({
            'success': True,
            'files': file_info
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/preview-sample', methods=['POST'])
def preview_sample():
    """Preview raw data from sample file before processing"""
    try:
        data = request.json
        sample_name = data.get('sample_name')
        
        if not sample_name:
            return jsonify({
                'success': False,
                'error': 'Sample file name is required'
            }), 400
        
        # Get sample file path
        sample_dir = os.path.join(PROJECT_ROOT, 'data', 'input')
        file_path = os.path.join(sample_dir, sample_name)
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': 'Sample file not found'
            }), 404
        
        # Read raw data from Excel file (first 20 rows)
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({
                'success': False,
                'error': 'File is empty'
            }), 400
        
        headers = rows[0]
        max_preview_rows = 20
        
        # Convert rows to list of dicts
        preview_data = []
        for row in rows[1:max_preview_rows+1]:
            preview_data.append(dict(zip(headers, row)))
        
        total_rows = len(rows) - 1  # Exclude header
        
        wb.close()
        
        return jsonify({
            'success': True,
            'preview_data': preview_data,
            'total_rows': total_rows,
            'showing_rows': len(preview_data),
            'headers': list(headers)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle file upload and process"""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': 'Only .xlsx files are allowed'
            }), 400
        
        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(file_path)
        
        # Get orchestrator (lazy initialization)
        orch = get_orchestrator()
        if orch is None:
            return jsonify({
                'success': False,
                'error': 'Orchestrator not initialized. Check server logs.'
            }), 500
        
        # Process file
        job_id = str(uuid.uuid4())
        summary, output_files = orch.run(file_path)
        
        # Get preview data
        reconciliation_file_path = output_files.get('general_reconciliation')
        print(f"DEBUG: Getting preview data from: {reconciliation_file_path}")
        preview_data = get_preview_data(reconciliation_file_path)
        print(f"DEBUG: Preview data loaded: {len(preview_data)} rows")
        
        # Store results
        results_store[job_id] = {
            'summary': summary.to_dict(),
            'output_files': output_files,
            'preview_data': preview_data,
            'timestamp': datetime.now().isoformat()
        }
        print(f"DEBUG: Stored results for job {job_id}, preview_data length: {len(preview_data)}")
        
        # Store job_id in session
        session['job_id'] = job_id
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'summary': summary.to_dict(),
            'output_files': output_files
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/process-sample', methods=['POST'])
def process_sample():
    """Process a sample file"""
    try:
        data = request.get_json()
        sample_name = data.get('sample_name')
        
        if not sample_name:
            return jsonify({
                'success': False,
                'error': 'No sample file specified'
            }), 400
        
        file_path = os.path.join(PROJECT_ROOT, 'data', 'input', sample_name)
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': f'Sample file not found: {sample_name}'
            }), 404
        
        # Get orchestrator (lazy initialization)
        orch = get_orchestrator()
        if orch is None:
            return jsonify({
                'success': False,
                'error': 'Orchestrator not initialized. Check server logs.'
            }), 500
        
        # Process file
        job_id = str(uuid.uuid4())
        summary, output_files = orch.run(file_path)
        
        # Get preview data
        reconciliation_file_path = output_files.get('general_reconciliation')
        print(f"DEBUG: Getting preview data from: {reconciliation_file_path}")
        preview_data = get_preview_data(reconciliation_file_path)
        print(f"DEBUG: Preview data loaded: {len(preview_data)} rows")
        
        # Store results
        results_store[job_id] = {
            'summary': summary.to_dict(),
            'output_files': output_files,
            'preview_data': preview_data,
            'timestamp': datetime.now().isoformat()
        }
        print(f"DEBUG: Stored results for job {job_id}, preview_data length: {len(preview_data)}")
        
        # Store job_id in session
        session['job_id'] = job_id
        
        return jsonify({
            'success': True,
            'job_id': job_id,
            'summary': summary.to_dict(),
            'output_files': output_files
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/results/<job_id>', methods=['GET'])
def get_results(job_id):
    """Get results for a job"""
    if job_id not in results_store:
        return jsonify({
            'success': False,
            'error': 'Job not found'
        }), 404
    
    return jsonify({
        'success': True,
        'results': results_store[job_id]
    })


@app.route('/api/download/<file_type>', methods=['GET'])
def download_file(file_type):
    """Download output file"""
    try:
        # Try to get job_id from query parameter first, then session
        job_id = request.args.get('job_id') or session.get('job_id')
        
        if not job_id or job_id not in results_store:
            # If no session, try to use the latest job
            if results_store:
                job_id = list(results_store.keys())[-1]
            else:
                return jsonify({
                    'success': False,
                    'error': 'No results available'
                }), 404
        
        results = results_store[job_id]
        output_files = results.get('output_files', {})
        
        if file_type == 'reconciliation':
            file_path = output_files.get('general_reconciliation')
        elif file_type == 'master_missing':
            file_path = output_files.get('master_missing')
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid file type'
            }), 400
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': 'File not found'
            }), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=os.path.basename(file_path)
        )
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/download-by-job/<job_id>/<file_type>', methods=['GET'])
def download_file_by_job(job_id, file_type):
    """Download output file by job ID"""
    try:
        if job_id not in results_store:
            return jsonify({
                'success': False,
                'error': 'Job not found'
            }), 404
        
        results = results_store[job_id]
        output_files = results.get('output_files', {})
        
        if file_type == 'reconciliation':
            file_path = output_files.get('general_reconciliation')
        elif file_type == 'master_missing':
            file_path = output_files.get('master_missing')
        else:
            return jsonify({
                'success': False,
                'error': 'Invalid file type'
            }), 400
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'error': 'File not found'
            }), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=os.path.basename(file_path)
        )
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/results')
def results_page():
    """Results display page"""
    job_id = session.get('job_id')
    
    if not job_id or job_id not in results_store:
        return render_template('results.html', error='No results available')
    
    results = results_store[job_id]
    return render_template('results.html', results=results)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
